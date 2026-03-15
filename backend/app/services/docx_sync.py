import io
import re
from typing import List, Optional
from pydantic import BaseModel
from docx import Document
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

class StructuredItem(BaseModel):
    category: str
    finding: str
    action: str
    owner: str
    due_date: str
    priority: str
    risk: str
    control: str
    status: str

class DocxSyncService:
    def parse_narrative_to_json(self, docx_content: bytes) -> List[StructuredItem]:
        doc = Document(io.BytesIO(docx_content))
        items = []
        current_category = "General"
        
        # State machine for extraction
        current_item = None

        patterns = {
            'finding': [r'Finding\s*:\s*(.*)', r'Observation\s*:\s*(.*)'],
            'action': [r'Action\s*:\s*(.*)', r'Recommendation\s*:\s*(.*)'],
            'owner': [r'Owner\s*:\s*(.*)'],
            'due_date': [r'Due Date\s*:\s*(.*)', r'Deadline\s*:\s*(.*)'],
            'priority': [r'Priority\s*:\s*(.*)'],
            'risk': [r'Risk\s*:\s*(.*)', r'Impact\s*:\s*(.*)'],
            'control': [r'Control\s*:\s*(.*)', r'Mitigation\s*:\s*(.*)'],
            'status': [r'Status\s*:\s*(.*)']
        }

        def finalize_item(item):
            if item and (item['finding'].strip() or item['action'].strip()):
                items.append(StructuredItem(**item))

        for para in doc.paragraphs:
            text = para.text.strip()
            if not text:
                continue

            if para.style.name.startswith('Heading'):
                finalize_item(current_item)
                current_category = text
                current_item = None
                continue

            # Detect start of a new finding/block
            is_start = any(re.search(rf'^{label}\s*:', text, re.I) for label in ['Finding', 'Observation'])
            
            if is_start:
                finalize_item(current_item)
                current_item = {k: "" for k in ['finding', 'action', 'owner', 'due_date', 'priority', 'risk', 'control', 'status']}
                current_item['category'] = current_category or "General"

            if current_item is not None:
                # Try to extract any field from this paragraph
                for key, regexes in patterns.items():
                    for reg in regexes:
                        match = re.search(reg, text, re.I)
                        if match:
                            val = match.group(1).strip()
                            # Handle multiple tags in one line (e.g., Owner: X | Priority: Y)
                            if "|" in val:
                                val = val.split("|")[0].strip()
                            current_item[key] = val
                            break
                
                # Check for pipe-separated metadata in the same paragraph
                if "|" in text:
                    parts = text.split("|")
                    for p in parts:
                        for key, regexes in patterns.items():
                            for reg in regexes:
                                m = re.search(reg, p.strip(), re.I)
                                if m:
                                    current_item[key] = m.group(1).strip()
                                    break

        finalize_item(current_item)
        return items

    def generate_excel_workbook(self, items: List[StructuredItem]) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "Transformation Data"
        orange_fill = PatternFill(start_color="FF6200", end_color="FF6200", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        center_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        headers = ["Category", "Finding", "Action Item", "Owner", "Due Date", "Priority", "Risk", "Control", "Status"]
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = orange_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = thin_border
        for item in items:
            ws.append([item.category, item.finding, item.action, item.owner, item.due_date, item.priority, item.risk, item.control, item.status])
        # Add zebra stripes and borders
        light_orange_fill = PatternFill(start_color="FFF2E6", end_color="FFF2E6", fill_type="solid")
        open_fill = PatternFill(start_color="FF6200", end_color="FF6200", fill_type="solid")
        closed_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
        white_font = Font(color="FFFFFF", bold=True)

        for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
            is_even = i % 2 == 0
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                if is_even:
                    cell.fill = light_orange_fill
                
                if cell.column == 9: # Status
                    val = str(cell.value).upper()
                    if "OPEN" in val:
                        cell.fill = open_fill
                        cell.font = white_font
                    elif "CLOSED" in val:
                        cell.fill = closed_fill
                        cell.font = white_font
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
            ws.column_dimensions[column].width = min(40, max_length + 2)
        ws.auto_filter.ref = ws.dimensions
        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

docx_sync_service = DocxSyncService()
